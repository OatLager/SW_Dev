import tkinter as tk
from tkinter import ttk, filedialog
import pandas as pd
import os
import func_OpenAI_api_v03 as api
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import openpyxl
import subprocess
import sys
import signal
import json

class ExcelOpenAIApp:
    def __init__(self, root_api):
        self.root_api = root_api
        self.api_key_var = tk.StringVar()
        self.json_template_file_var = tk.StringVar()
        self.excel_file_path_var = tk.StringVar()
        self.code_folder_path_var = tk.StringVar()
        self.progress_bar_var = tk.DoubleVar()
        self.assistant_id = None

        # Set default paths
        current_directory = self.get_exe_path()
        self.api_key_var.set(os.path.join(current_directory, "OpenAI_API_Key.txt"))
        self.json_template_file_var.set(os.path.join(current_directory, "function_template.json"))

        self.setup_ui()

    def cleanup(self, signum, frame):
        if self.assistant_id:
            api.delete_assistant(self.assistant_id)
        sys.exit(0) 

    def get_exe_path(self):
        """ 실행파일(.exe) 또는 스크립트(.py)의 실행 경로를 찾음 """
        if getattr(sys, 'frozen', False):  # PyInstaller 실행파일인지 확인
            return os.path.dirname(sys.executable)  # 실행파일 위치
        return os.path.dirname(os.path.abspath(__file__))  # 스크립트 위치

    def setup_ui(self):
        self.root_api.title("Excel OpenAI Processing")
        self.root_api.geometry("350x550")

        # 창을 화면 중앙에 위치시키기
        window_width = 350
        window_height = 550
        screen_width = self.root_api.winfo_screenwidth()
        screen_height = self.root_api.winfo_screenheight()
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        self.root_api.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

        # Section : OpenAI API Key  -----------------------------------------------------------------
        path_api_frame = ttk.LabelFrame(self.root_api, text="OpenAI API Key", padding="10")
        path_api_frame.pack(fill="x", pady=(0, 10))

        ttk.Entry(path_api_frame, textvariable=self.api_key_var).pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(path_api_frame, text="Browse", command=self.browse_api_key_var_file).pack(side="right")

        # Section : Output Template  -----------------------------------------------------------------
        path_template_frame = ttk.LabelFrame(self.root_api, text="output template", padding="10")
        path_template_frame.pack(fill="x", pady=(0, 10))
        ttk.Entry(path_template_frame, textvariable=self.json_template_file_var).pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(path_template_frame, text="Browse", command=self.browse_json_template_file).pack(side="right")

        # Section : Excel File  -----------------------------------------------------------------
        path_excel_frame = ttk.LabelFrame(self.root_api, text="Excel File", padding="10")
        path_excel_frame.pack(fill="both", expand=True, pady=(0, 10))

        # File Path Row 
        path_file_frame = ttk.Frame(path_excel_frame)
        path_file_frame.pack(fill="x", pady=(0, 5))

        ttk.Entry(path_file_frame, textvariable=self.excel_file_path_var).pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(path_file_frame, text="Browse", command=self.browse_file).pack(side="right")

        # Lists Frame -----------------------------------------------------------------
        lists_frame = ttk.Frame(path_excel_frame)
        lists_frame.pack(fill="both", expand=True)

        # File List 
        file_frame = ttk.LabelFrame(lists_frame, text="Source File List", padding="10")
        file_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        self.file_listbox = tk.Listbox(file_frame, selectmode=tk.MULTIPLE)
        self.file_listbox.pack(fill="both", expand=True)
        self.file_listbox.bind('<<ListboxSelect>>', self.on_listbox_select)

        # Select Frame with Checkboxes -----------------------------------------------------------------
        select_frame = ttk.LabelFrame(lists_frame, text="Select", padding="10")
        select_frame.pack(side="right", fill="y", padx=(5, 0))
        # all
        self.select_all_var = tk.BooleanVar()
        select_all_checkbox = ttk.Checkbutton(select_frame, text="All", variable=self.select_all_var, command=self.toggle_select_file)
        select_all_checkbox.pack(anchor="w")
        # cpp
        self.select_cpp_var = tk.BooleanVar()
        select_cpp_checkbox = ttk.Checkbutton(select_frame, text="cpp", variable=self.select_cpp_var, command=self.toggle_select_file)
        select_cpp_checkbox.pack(anchor="w")
        # c
        self.select_c_var = tk.BooleanVar()
        select_c_checkbox = ttk.Checkbutton(select_frame, text="c", variable=self.select_c_var, command=self.toggle_select_file)
        select_c_checkbox.pack(anchor="w")
        # h
        self.select_h_var = tk.BooleanVar()
        select_h_checkbox = ttk.Checkbutton(select_frame, text="h", variable=self.select_h_var, command=self.toggle_select_file)
        select_h_checkbox.pack(anchor="w")
        # hpp
        self.select_hpp_var = tk.BooleanVar()
        select_hpp_checkbox = ttk.Checkbutton(select_frame, text="hpp", variable=self.select_hpp_var, command=self.toggle_select_file)
        select_hpp_checkbox.pack(anchor="w")

        # Progress Bar and Start Button Frame
        start_frame = ttk.Frame(self.root_api, padding="10")
        start_frame.pack(fill="both", expand=True, pady=(0, 0))

        self.progress_bar = ttk.Progressbar(start_frame, variable=self.progress_bar_var, maximum=100)
        self.progress_bar.pack(side="left", fill="x", expand=True, padx=(0, 5))

        start_button = ttk.Button(start_frame, text="Start", command=self.start_process)
        start_button.pack(side="right", padx=(0, 5))

        bottom_frame = ttk.Frame(self.root_api, padding="10")
        bottom_frame.pack(fill="both", expand=True)
        open_button = ttk.Button(bottom_frame, text="Open", command=self.open_output_folder)
        open_button.grid(row=0, column=2, padx=(5, 0), pady=(0, 5), sticky="e")
        json_button = ttk.Button(bottom_frame, text="json->xls", command=self.json_to_excel)
        json_button.grid(row=0, column=0, padx=(0, 5), pady=(0, 5), sticky="w")
        flow_button = ttk.Button(bottom_frame, text="flowchart", command=self.flowchart_to_excel)
        flow_button.grid(row=0, column=1, padx=(10, 10), pady=(0, 5))
        
    def open_output_folder(self):
        if os.path.exists(self.excel_file_path_var.get()):
            try:
                if os.path.exists(self.excel_file_path_var.get()):
                    if os.name == 'nt':
                        os.startfile(self.excel_file_path_var.get())
                    elif os.name == 'posix':
                        subprocess.Popen(['xdg-open', self.excel_file_path_var.get()])
                else:
                    subprocess.Popen(['xdg-open', self.excel_file_path_var.get()])
            except FileNotFoundError:
                tk.messagebox.showError("Error", f"Folder not found: {self.excel_file_path_var.get()}")
        else:
            tk.messagebox.showError("Error", "No folder path found.")

    def on_listbox_select(self, event):
        self.select_all_var.set(False)
        self.select_cpp_var.set(False)
        self.select_c_var.set(False)
        self.select_h_var.set(False)
        self.select_hpp_var.set(False)

    def toggle_select_file(self):
        if self.select_all_var.get():
            self.file_listbox.select_set(0, tk.END)
            self.select_cpp_var.set(False)
            self.select_c_var.set(False)
            self.select_h_var.set(False)
            self.select_hpp_var.set(False)
        else:
            self.file_listbox.select_clear(0, tk.END)
            for i in range(self.file_listbox.size()):
                item = self.file_listbox.get(i)
                if self.select_cpp_var.get() and item.endswith('.cpp'):
                    self.file_listbox.select_set(i)
                if self.select_c_var.get() and item.endswith('.c'):
                    self.file_listbox.select_set(i)
                if self.select_h_var.get() and item.endswith('.h'):
                    self.file_listbox.select_set(i)
                if self.select_hpp_var.get() and item.endswith('.hpp'):
                    self.file_listbox.select_set(i)

    def browse_api_key_var_file(self):
        api_key_var_file = filedialog.askopenfilename(
            filetypes=[("OpenAI API Key files(.txt)", "*.txt")]
        )
        
        if api_key_var_file:
            print(f'  🔹 API Key Path : {api_key_var_file}')
            self.api_key_var.set(api_key_var_file)
            self.root_api.update_idletasks()

    def browse_json_template_file(self):
        json_template_file = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")]
        )
        self.json_template_file_var.set(json_template_file)
        print(f'  🔹 JSON Template File : {json_template_file}')
        
    def browse_file(self):
        excel_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if excel_file:
            self.excel_file_path_var.set(excel_file)
            print(f'  🔹 Excel File : {excel_file}')
            self.load_excel_info()

    def load_excel_info(self):
        try:
            # Clear existing items
            self.file_listbox.delete(0, tk.END)

            # Load Excel file
            excel_file = pd.ExcelFile(self.excel_file_path_var.get())
            
            # Excel 파일의 첫 번째 시트의 G4 셀[Path]을 확인하여 소스코드 경로 H4셀 확인.
            first_sheet = excel_file.sheet_names[0]
            sheet_data = excel_file.parse(first_sheet)
            if sheet_data.iat[2, 6] == 'Path':  # 'G' is the 7th column, so index is 6
                code_folder_path = sheet_data.iat[2, 7]  # 'H' is the 8th column, so index is 7

            # Excel 파일의 두 번째 시트부터 시트 이름을 리스트박스에 추가(시트 이름:소스코드 파일 이름)
            # Add sheet names to listbox, starting from the second sheet
            for sheet in excel_file.sheet_names[1:]:
                self.file_listbox.insert(tk.END, sheet)
            
            # Allow MULTIPLE selection for sheets
            self.file_listbox.config(selectmode=tk.MULTIPLE)

            # return code_folder_path
            self.code_folder_path_var.set(code_folder_path)

        except Exception as e:
            print(f"❌ Error - Loading Excel file: {str(e)}")
            # return None

    def get_selected_files(self, code_folder_path, selected_files):
        existing_files = []
        if os.path.exists(code_folder_path):
            for root_api, _, files in os.walk(code_folder_path):
                for file in selected_files:
                    if file in files:
                        file_path = os.path.join(root_api, file)
                        existing_files.append(file_path)
                        print(f"  🔹 Code File found: {file_path}")
            missing_files = set(selected_files) - set(os.path.basename(f) for f in existing_files)
            for file in missing_files:
                print(f"❌ Error - Code File not found: {file}")
        else:
            print(f"❌ Error - Invalid Code File Path: {code_folder_path}")

    def json_to_excel(self):
        json_file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")]
        )
        excel_file_path_var = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        sheet_name = os.path.basename(json_file_path).replace('.json', '')
        print(f'  🔹 JSON File : {json_file_path}')
        print(f'  🔹 Excel File : {excel_file_path_var}')
        print(f'  🔹 Sheet Name : {sheet_name}')
        json_function_list = []
        
        if os.path.exists(json_file_path):
            with open(json_file_path, "r", encoding="utf-8") as json_file:
                try:
                    existing_data = json.load(json_file)  # 기존 데이터 로드
                except json.JSONDecodeError:
                    print("❌ Error: JSON 파일이 손상됨. 빈 데이터로 초기화합니다.")
                    existing_data = {"json_function_list": []}
        else:
            print("⚠️ Warning: JSON 파일이 존재하지 않음. 새로 만듭니다.")
            existing_data = {"json_function_list": []}

        # 2️⃣ 데이터 가져오기
        if isinstance(existing_data, dict) and "functions" in existing_data and isinstance(existing_data["functions"], list):
            classes = [func["class"] for func in existing_data["functions"]]
            names = [func["name"] for func in existing_data["functions"]]
            summaries = [func["summary"] for func in existing_data["functions"]]
            flows = [func["flow"] for func in existing_data["functions"]]
        else:
            print("❌ Error: JSON 구조가 올바르지 않습니다.")
            classes, names, summaries, flows = [], [], [], []

        json_function_list.append((classes, names, summaries, flows))

        # Excel 파일 업데이트
        self.update_excel_with_json(json_function_list, excel_file_path_var, sheet_name)

    def flowchart_to_excel(self):
        # open ai api key 설정
        api_key_var_file = self.api_key_var.get()
        if api_key_var_file:
            with open(api_key_var_file, 'r') as file:
                my_api_key = file.read().strip()

        code_folder_path = self.code_folder_path_var.get()
        print(f"may api key : {my_api_key}")
        print(f"  🔹 Code Folder Path : {code_folder_path}")
        print(f"  🔹 API Key : {my_api_key}")
                                
        if not my_api_key:
            print("  ⚠️  Warning - No API key found.")
            return
        if not code_folder_path:
            print("  ⚠️  Warning - No code folder path found.")
            return
        api.set_openai_api_key(api_key_var_file)

        excel_file_path = self.excel_file_path_var.get()
        print(f"  🔹 Excel File Path : {excel_file_path}")

        # Get selected files before clearing the listbox
        selected_files = [self.file_listbox.get(i) for i in self.file_listbox.curselection()]
        if not selected_files:
            print("  ⚠️  Warning - No Code files selected.")
            return
        else:
            print(f"  🔹 Selected code files: {selected_files}")

        for sheet in selected_files:
            count = f'({selected_files.index(sheet) + 1}/{len(selected_files)})'
            self.progress_bar_var.set(len(selected_files) * (selected_files.index(sheet) + 1))
            self.root_api.update()
            print(f"\n🔸 {count}Processing - {sheet} flowchart------------------------------------------------------")
            self.write_flowchart(excel_file_path, sheet)
            
        self.progress_bar_var.set(100)
        self.root_api.update()
        print("\n\n✅ Success - Processing completed.\n\n")

    def write_to_excel(self, code_folder_path, excel_file_path_var, sheet_name):
        excel_function_list = []
        try:
            # Load the Excel file and the specified sheet
            df = pd.read_excel(excel_file_path_var, sheet_name=sheet_name)
            
            # Find rows where 'Category' column contains 'definition'
            definition_rows = df[df['Category'].str.contains(' (definition)', na=False, regex=False)]
            
            # Extract the 'Name' column from these rows
            excel_function_list = definition_rows[['Field', 'Name']].values.tolist()
            if excel_function_list:
                message_list = []
                for field, name in excel_function_list:
                    if pd.isna(field): # NaN 처리
                        field = ''
                    message_list.append(f"{field}의 {name}")
                # print(f"  🔹 Message_list : \n       {message_list}")
                code_file_path = os.path.join(code_folder_path, sheet_name)

                # api 실행 및 json 파일, function list 생성 [[names], [summaries], [sequences]]
                json_function_list = self.list_of_api_call(excel_file_path_var, code_file_path, message_list)
                # Excel 파일 업데이트
                self.update_excel_with_json(json_function_list, excel_file_path_var, sheet_name)
            else:
                print("⚠️  Warning - No 'Function (definition)' found.")
        except Exception as e:
            print(f"❌ Error - Updating Excel file: \n{str(e)}")
        
        # File flowchart
        print("\n\n  Step(5/5). OpenAI API - Flowdiagram 실행")
        self.write_flowchart(excel_file_path_var, sheet_name)

    # File flowchart 생성 요청(api 실행) 및 엑셀 파일에 이미지 삽입
    def write_flowchart(self, excel_file_path, sheet_name):
        if sheet_name.endswith(('.cpp', '.c')):
            # Excel 파일 경로에 flowchart 폴더 생성
            output_dir = os.path.dirname(excel_file_path)
            # "flowchart" 폴더 경로 생성
            flowchart_dir = os.path.join(output_dir, "flowchart")
            # 폴더가 없으면 생성
            os.makedirs(flowchart_dir, exist_ok=True)

            code_file_path = os.path.join(self.code_folder_path_var.get(), sheet_name)
            print(f"  🔹 Code File Path : {code_file_path} / write_flowchart")
            mermaid_flowchart = api.execute_openai_assistant_to_create_flowchart(code_file_path)
            # write mermaid_flowchart to mmd file
            mermaid_flowchart_mmd = os.path.join(flowchart_dir, sheet_name + ".mmd")
            mermaid_png = os.path.join(flowchart_dir, sheet_name + ".png")

            with open(mermaid_flowchart_mmd, 'w', encoding='utf-8') as f:
                f.write(mermaid_flowchart)
            subprocess.run(["mmdc", "-i", mermaid_flowchart_mmd, "-o", mermaid_png], shell=True)
            # insert image to excel file
            self.insert_image_to_excel(excel_file_path, sheet_name, mermaid_png)
        else:
            print("⚠️  Warning - Flowchart Only supported for source file(c, cpp).")

    # 함수 : 엑셀 파일에 생성한 flowchart 이미지 삽입 [A15]
    def insert_image_to_excel(self, excel_file_path, sheet_name, image_file_path):
        try:
            wb_existing = load_workbook(excel_file_path)
            if sheet_name in wb_existing.sheetnames:
                ws = wb_existing[sheet_name]
                flowchart_image = openpyxl.drawing.image.Image(image_file_path)
                flowchart_image.anchor = 'A15'
                ws.add_image(flowchart_image)
                wb_existing.save(excel_file_path)
                print(f"  🔸 Excel 파일 업데이트 완료 : {excel_file_path}\n")
            else:
                print(f"❌ Error - Excel 시트 확인 불가 : {wb_existing.sheetnames}")
        except Exception as e:
            print(f"❌ Error - Inserting image to Excel file: \n{str(e)}\n")

    def list_of_api_call(self, excel_file_path_var, code_file_path, excel_function_list):
        print("\n  Step(1/5) : OpenAI API - Input Data 설정")
        print(f"  🔹 code_file_path : {code_file_path}")
        print(f"  🔹 excel_file_path_var : {excel_file_path_var}")
        print(f"  🔹 excel_function_list : {excel_function_list}")

        # json 파일 저장 경로 및 이름 설정(Excel파일과 동일 경로, 코드파일 이름으로 저장)
        json_template_file_path = self.json_template_file_var.get() # r"C:\Users\ypelec\Desktop\SW_Dev\code_analysis\function_template.json" # json 템플릿 파일 경로
        json_file_path = os.path.dirname(excel_file_path_var)  # json 파일 생성 경로
        json_file_name = os.path.basename(code_file_path)  # json 파일 생성 이름(code file 확장자 포함)

        # (!api 파일첨부에 헤더파일 불가-> 소스파일로 변경).h or .hpp 파일을 .cpp 파일로 변환 후 저장
        if code_file_path.endswith('.h') or code_file_path.endswith('.hpp'):
            # Create 'for_API_Analysis' folder if it doesn't exist
            analysis_folder = os.path.join(os.path.dirname(code_file_path), 'API_Analysis_for_Header')
            os.makedirs(analysis_folder, exist_ok=True)
            
            # 파일 변환 .h or .hpp file -> .cpp file
            # Save the new .cpp file in the 'API_Analysis_for_Header' folder
            if code_file_path.endswith('.h'):
                new_code_file_path = os.path.join(analysis_folder, os.path.basename(code_file_path).replace('.h', '.cpp'))
            elif code_file_path.endswith('.hpp'):
                new_code_file_path = os.path.join(analysis_folder, os.path.basename(code_file_path).replace('.hpp', '.cpp'))
            # Read the .h or .hpp file and write the content to the new .cpp file
            with open(code_file_path, 'r') as header_file:
                content = header_file.read()
            with open(new_code_file_path, 'w') as cpp_file:
                cpp_file.write(content)
            print(f"     - Converted {os.path.basename(code_file_path)} to .cpp and saved to: {new_code_file_path}")
            # Update the code_file_path to the new .cpp file path
            code_file_path = new_code_file_path
            
        # API 요청 메시지 생성, 코드 파일 내의 함수 리스트에 대한 설명 요청
        if excel_function_list:
            message_content = []
            for excel_function in excel_function_list:
                message_content.append(f"{excel_function} 함수 설명")         
        
        # API 실행
        print("\n  Step(2/5). OpenAI API - 함수 분석 실행")
        API_response_text, self.assistant_id, code_file_id, json_template_id = api.execute_openai_assistant(message_content, code_file_path, json_template_file_path)
       # Assistant 삭제
        api.delete_upload_file(code_file_id)
        api.delete_upload_file(json_template_id)
        api.delete_assistant(self.assistant_id)

        # API 응답 JSON 파일 저장
        print("\n  Step(3/5). OpenAI API - 함수 분석 실행 결과 저장(JSON)")
        json_function_list = api.response_save_json(API_response_text, json_file_path, json_file_name)
        return json_function_list

    def update_excel_with_json(self, content, excel_file_path_var, sheet_name):
        print("\n  Step(4/5). OpenAI API - 함수 분석 Excel 파일 업데이트")
        if content and excel_file_path_var:
            wb_existing = load_workbook(excel_file_path_var)

            if sheet_name in wb_existing.sheetnames:
                print(f"  🔹 Excel sheet 확인 : {sheet_name}")
                ws = wb_existing[sheet_name]
                table_headers = [cell.value for cell in ws[1]]
                if "Name" in table_headers:
                    name_col_index = table_headers.index("Name") + 1
                    summary_col_index = table_headers.index("Summary") + 1 if "Summary" in table_headers else None
                    description_col_index = table_headers.index("Description") + 1 if "Description" in table_headers else None

                    count = len(content[0][0])  # 전체 함수 개수 (json 파일 내의 class 갯수로 확인)
                    processed_count = 0  # 처리된 함수 개수
                    # Excel 파일에 JSON 파일 내용 추가 - name과 class가 동일하면 summary, sequence 업데이트
                    for classes, names, summaries, flows in content:
                        for j, (class_name, name) in enumerate(zip(classes, names)):  # class와 name을 동시에 사용
                            for row in ws.iter_rows(min_row=2, max_col=name_col_index, max_row=ws.max_row):
                                for cell in row:
                                    if cell.value == name:
                                        # 해당 name이 일치하는 경우, class도 확인(class가 없는 경우는 패스)
                                        if class_name:
                                            if cell.offset(0, -3).value != class_name:  # name의 세 칸 왼쪽 셀과 비교
                                                continue
                                        processed_count += 1       
                                        print(f"  🔹 - Function definition : {class_name}의 {name} 함수 확인.")        
                                        if summary_col_index:
                                            ws.cell(row=cell.row, column=summary_col_index, value=summaries[j])
                                        if description_col_index:
                                            cell = ws.cell(row=cell.row, column=description_col_index, value="\n".join(flows[j]))
                                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(vertical='top')

                    wb_existing.save(excel_file_path_var)
                    if processed_count == count:
                        print(f"  🔸 {processed_count}/{count} 함수 업데이트 확인")
                    else:
                        print(f"⚠️  Warning - {processed_count}/{count} 함수 업데이트 확인")
                    print(f"  🔸 Excel 파일 업데이트 완료 : {excel_file_path_var}")

                else:
                    print("❌ Error - Excel 시트 테이블 : 'Name' 항목 확인 불가.")
            else:
                print(f"❌ Error - Excel 시트 확인 불가 : {wb_existing.sheetnames}")
        else:
            print("❌ Error - Excel 파일 저장 경로 확인 불가.")

    def start_process(self):
        print("\n[ Processing started... ] ==================================================================================")
        
        api_key_var_file = self.api_key_var.get()
        if api_key_var_file:
            with open(api_key_var_file, 'r') as file:
                my_api_key = file.read().strip()

        code_folder_path = self.code_folder_path_var.get()
        print(f"may api key : {my_api_key}")
        print(f"  🔹 Code Folder Path : {code_folder_path}")
        print(f"  🔹 API Key : {my_api_key}")
                                
        if not my_api_key:
            print("  ⚠️  Warning - No API key found.")
            return
        if not code_folder_path:
            print("  ⚠️  Warning - No code folder path found.")
            return
        api.set_openai_api_key(api_key_var_file)

        # Get selected files before clearing the listbox
        selected_files = [self.file_listbox.get(i) for i in self.file_listbox.curselection()]
        if not selected_files:
            print("  ⚠️  Warning - No Code files selected.")
            return
        else:
            print(f"  🔹 Selected code files: {selected_files}")

        # Add your processing logic here
        self.progress_bar_var.set(0)
        
        # Simulate progress
        self.get_selected_files(code_folder_path, selected_files)
        self.progress_bar_var.set(10)
        self.root_api.update()
        excel_file_path = self.excel_file_path_var.get()
        for sheet in selected_files:
            count = f'({selected_files.index(sheet) + 1}/{len(selected_files)})'
            self.progress_bar_var.set(10 + (80 / len(selected_files) * (selected_files.index(sheet) + 1)))
            self.root_api.update()
            print(f"\n🔸 {count}Processing - {sheet} --------------------------------------------------------------")
            self.write_to_excel(code_folder_path, excel_file_path, sheet)

        self.progress_bar_var.set(100)
        self.root_api.update()
        print("\n\n✅ Success - Processing completed.\n\n")

        def on_closing():
            # self.root_api.destroy()
            if self.assistant_id:
                api.delete_assistant(self.assistant_id)
            os._exit(0)  # 강제 종료

        self.root_api.protocol("WM_DELETE_WINDOW", on_closing)

def main():
    signal.signal(signal.SIGINT, ExcelOpenAIApp.cleanup)
    root_api = tk.Toplevel()  # Toplevel을 사용하여 서브 윈도우로 생성
    app = ExcelOpenAIApp(root_api)

    root_api.mainloop()

if __name__ == "__main__":
    main()