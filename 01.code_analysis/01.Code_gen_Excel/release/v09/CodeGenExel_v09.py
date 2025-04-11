# 네임스페이스 안에 함수 추출해야됨. 


import os
import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess
import func_extract_code_v02 
import convert_compile_cmd as ccc
import OpenAI_api_v07

# ====================================================================================================
# output 파일 생성(xlsx)
def generate_output_filename(output_folder, base_folder):
    date_str = datetime.now().strftime("%Y%m%d")
    base_name = os.path.basename(base_folder)
    output_filename = f"{base_name}_{date_str}.xlsx"
    return os.path.join(output_folder, output_filename)
# source code 폴더 선택
def select_input_folder():
    folder_path = filedialog.askdirectory()
    folder_path_var.set(folder_path)
# output folder 경로 선택
def select_output_folder():
    output_folder = filedialog.askdirectory()
    output_folder_var.set(output_folder)
# output folder 열기
def open_output_folder():
    if os.path.exists(output_folder_var.get()):
        try:
            if os.path.exists(output_folder_var.get()):
                if os.name == 'nt':
                    os.startfile(output_folder_var.get())
                elif os.name == 'posix':
                    subprocess.Popen(['xdg-open', output_folder_var.get()])
            else:
                subprocess.Popen(['xdg-open', output_folder_var.get()])
        except FileNotFoundError:
            tk.messagebox.showerror("Error", f"Folder not found: {output_folder_var.get()}")
# compile_commands.json 파일 선택
def select_compile_commands():
    file_path = filedialog.askopenfilename(title="Select the compile_commands.json file", filetypes=[("JSON files", "*.json")])

    if not file_path:
        raise ValueError("No file selected")
    else:
        compile_commands_var.set(file_path)
        return file_path
def convert_compile_commands():
    path_compile_commands = ccc.convert_compile_commands()
    if path_compile_commands:
        compile_commands_var.set(path_compile_commands)
    
def api_call():
    OpenAI_api_v07.main()

# ====================================================================================================
# 엑셀 파일 작성
def write_to_excel(file_data, output_file, folder_path, header_files, source_files):
    workbook = Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # 1. 파일 목록 시트 작성 -------------------------------------------------------------
    sheet = workbook.active
    sheet.title = "File List"
    header = ["No", "File Name", "Usage", "Basis", "Description"]
    sheet.append(header)
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    # 파일 목록 작성
    file_no = 1
    file_colors = {
        ".h": "FFFF99",
        ".hpp": "FFFF99",
        ".c": "FFFF99",
        ".cpp": "FFFF99"
    }
    for file in os.listdir(folder_path):
        if os.path.isfile(os.path.join(folder_path, file)) and file != os.path.basename(output_file):
            ext = os.path.splitext(file)[1]
            row = [file_no, file]  # No subdirectory, so second column is empty
            sheet.append(row)
            for cell in sheet[file_no + 1]:  # file_no + 1 because header is in the first row
                if ext in file_colors:
                    cell.fill = PatternFill(start_color=file_colors[ext], end_color=file_colors[ext], fill_type="solid")
            file_no += 1

    column_widths = [5, 30, 10, 20, 20]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # 추가 정보 작성(날짜, 시간, 폴더명, 경로, 생성프로그램, 추출파일 분류)
    sheet["G1"] = "Creation Date"
    sheet["H1"] = datetime.now().strftime("%Y-%m-%d")
    sheet["G2"] = "Creation Time"
    sheet["H2"] = datetime.now().strftime("%H:%M:%S")
    sheet["G3"] = "Selected Folder"
    sheet["H3"] = os.path.basename(folder_path)
    sheet["G4"] = "Path"
    sheet["H4"] = folder_path
    sheet["G5"] = "Generator"
    sheet["H5"] = os.path.basename(__file__)
    sheet["G6"] = "Extracted Files"
    sheet["H6"] = f".h: {len(header_files)} / .hpp: {len([file for file in header_files if file.endswith('.hpp')])} / .c: {len([file for file in source_files if file.endswith('.c')])} / .cpp: {len([file for file in source_files if file.endswith('.cpp')])}"


    for cell in ["G1", "G2", "G3", "G4", "G5", "G6"]:
        sheet[cell].fill = header_fill
        sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
        sheet[cell].font = Font(bold=True)


    # 파일 목록과 추가 정보에 굵은 바깥쪽 테두리 적용
    for row in sheet.iter_rows(min_row=1, max_row=file_no, min_col=1, max_col=len(header)):
        for cell in row:
            if cell.row == 1:
                cell.border = Border(top=thick_border.top, bottom=thin_border.bottom, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)
            elif cell.row == file_no:
                cell.border = Border(top=None, bottom=thick_border.bottom, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)
            else:
                cell.border = Border(top=None, bottom=thin_border.bottom, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)

    for row in sheet.iter_rows(min_row=1, max_row=6, min_col=len(header)+2, max_col=len(header)+3):
        for cell in row:
            if cell.row == 1:
                cell.border = Border(top=thick_border.top, bottom=thin_border.bottom, left=thick_border.left if cell.col_idx == len(header)+2 else thin_border.left, right=thick_border.right if cell.col_idx == len(header)+3 else thin_border.right)
            elif cell.row == 6:
                cell.border = Border(top=None, bottom=thick_border.bottom, left=thick_border.left if cell.col_idx == len(header)+2 else thin_border.left, right=thick_border.right if cell.col_idx == len(header)+3 else thin_border.right)
            else:
                cell.border = Border(top=None, bottom=thin_border.bottom, left=thick_border.left if cell.col_idx == len(header)+2 else thin_border.left, right=thick_border.right if cell.col_idx == len(header)+3 else thin_border.right)

    # 추가 정보 열 너비 설정
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 80

    # 2.파일 별 소스코드 추출 시트 작성  ----------------------------------------------------------------
    category_colors = {
        "Function": "FFFF99",
        "param": "FFFF55",
        "Constructor": "FFCC99",
        "Destructor": "FF9999",
        "Variable": "99CCFF",
        "Enum": "99FF99"
    }
    # 엑셀 시트 생성 (엑셀 허용 이름 최대 31자 제한)
    for file_name, (functions, constructors, destructors, variables, enums, param) in file_data.items():
        sheet_title = os.path.basename(file_name)[:31]  # Truncate title to 31 characters
        if sheet_title in workbook.sheetnames:
            base_title = sheet_title[:28]  # Reserve space for suffix
            suffix = 1
            while sheet_title in workbook.sheetnames:
                sheet_title = f"{base_title}_{suffix}"
                suffix += 1
        sheet = workbook.create_sheet(title=sheet_title)
        header = ["No", "File Name", "Line", "Access", "Specifiers", "Field", "Category", "Type", "Name", "Input", "Summary", "Description"]
        sheet.append(header)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        current_row = 2  # Start from the second row since the first row is the header
        for category, items in [("Variable", variables), ("Function", functions), ("Constructor", constructors), ("Destructor", destructors), ("Enum", enums), ("param", param)]:
            row_no = 1
            for item in items:
                row = [row_no] + list(item) + [""] + [""]  # Add an empty column for the Summary & Description
                sheet.append(row)
                for cell in sheet[current_row]:  # Use current_row to access the correct row
                    cell.fill = PatternFill(start_color=category_colors[category], end_color=category_colors[category], fill_type="solid")
                row_no += 1
                current_row += 1

        # 테이블 생성
        table_name = f"Table_{sheet_title}"
        tab = Table(displayName=table_name, ref=f"A1:L{sheet.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)
        sheet.freeze_panes = sheet['A2']

        # 스타일 설정(테두리 및 셀 크기)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(header)):
            for cell in row:
                if cell.row == 2:
                    cell.border = Border(top=thick_border.top, bottom=None, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)
                elif cell.row == sheet.max_row:
                    cell.border = Border(top=None, bottom=thick_border.bottom, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)
                else:
                    cell.border = Border(top=None, bottom=None, left=thick_border.left if cell.col_idx == 1 else thin_border.left, right=thick_border.right if cell.col_idx == len(header) else thin_border.right)
        column_widths = [5, 15, 5, 10, 18, 40, 18, 30, 40, 140, 70, 110]
        # ["No", "File Name", "Line", "Access", "Category", "Type", "Name", "Input", "Remarks", "Summary", "Description"]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = width

    workbook.save(output_file)

# ====================================================================================================
# 메인 실행 함수
def start_processing():
    # 컴파일 커맨드 파일 경로 선택 (옵션)
    path_compile_commands = compile_commands_var.get()
        
    # 폴더 경로 확인 ---------------------------------------------------------------
    folder_path = folder_path_var.get()
    folder_path = os.path.normpath(folder_path)
    output_folder = output_folder_var.get()
    if not folder_path or not output_folder:
        tk.messagebox.showerror("Error", "Please select both input and output folders.")
        return

    output_excel_path = generate_output_filename(output_folder, folder_path)

    # 파일 목록 추출 ---------------------------------------------------------------
    file_data = {}
    header_files = []
    source_files = []

    for file in os.listdir(folder_path):
        if file.endswith(".h") or file.endswith(".hpp"):
            file_path = os.path.join(folder_path, file)
            file_path = os.path.normpath(file_path)
            header_files.append(os.path.join(folder_path, file))
        elif file.endswith(".c") or file.endswith(".cpp"):
            file_path = os.path.join(folder_path, file)
            file_path = os.path.normpath(file_path)
            source_files.append(os.path.join(folder_path, file))



    total_files = len(header_files) + len(source_files)
    processed_files = 0

    progress_label.config(text="Extracting Source Code...")
    progress_bar['value'] = 0
    file_label.config(text="")

    open_button.config(state=tk.DISABLED)
    open_button.pack_forget()  # Hide the button initially

    root.update()

    # 파일별 소스코드 추출 ----------------------------------------------------------
    # 헤더 파일 코드 추출
    for header_file in header_files:

        # 진행상태 표시
        processed_files += 1
        progress_bar['value'] = (processed_files/(total_files+1))*100
        file_label.config(text=f"Processing {processed_files}/{total_files}\n{os.path.basename(header_file)}", anchor='w')
        root.update()
        print(f"\n\n🔷 [ Processing ({processed_files}/{total_files}) files - {os.path.basename(header_file)} ]")

        cursor = func_extract_code_v02.parse_project(header_file, path_compile_commands)
        functions, constructors, destructors, variables, enums, param = func_extract_code_v02.extract_source_code(cursor, header_file)
        print(f"\n🔹 functions : {len(functions)} / constructors : {len(constructors)} / destructors : {len(destructors)} / variables : {len(variables)} / enums : {len(enums)} / param : {len(param)}")
        print("-"*120)
        if functions or constructors or destructors or variables or enums or param:
            file_data[os.path.basename(header_file)] = (functions, constructors, destructors, variables, enums, param)
        else:
            print(f"⚠️  Warning - No functions or variables found in {os.path.basename(header_file)}")
        
    # 소스 파일 코드 추출
    for source_file in source_files:

        # 진행상태 표시
        processed_files += 1
        progress_bar['value'] = (processed_files/(total_files+1))*100
        file_label.config(text=f"Processing {processed_files}/{total_files}\n{os.path.basename(source_file)}", anchor='w')
        root.update()
        print(f"\n\n🔷 [ Processing ({processed_files}/{total_files}) files - {os.path.basename(source_file)} ]")

        cursor = func_extract_code_v02.parse_project(source_file, path_compile_commands)
        functions, constructors, destructors, variables, enums, param = func_extract_code_v02.extract_source_code(cursor, source_file)
        print(f"\n🔹 functions : {len(functions)} / constructors : {len(constructors)} / destructors : {len(destructors)} / variables : {len(variables)} / enums : {len(enums)} / param : {len(param)}")
        print("-"*120)
        if functions or constructors or destructors or variables or enums or param:
            file_data[os.path.basename(source_file)] = (functions, constructors, destructors, variables, enums, param)
        else:
            print(f"⚠️  Warning  - No functions or variables found in {os.path.basename(source_file)}")
    
    # 엑셀 파일 작성 ----------------------------------------------------------------
    write_to_excel(file_data, output_excel_path, folder_path, header_files, source_files)
    print(f"\n\n✅ Success - Functions and variables have been written to {output_excel_path}\n\n")

    # 진행상태 표시
    progress_bar['value'] = 100

    if progress_bar['value'] == 100:
        progress_label.config(text="Processing complete!")
        open_button.config(state=tk.NORMAL)
        progress_bar.pack_forget()
        open_button.pack()  # Show the button after processing is complete
        root.update()

    def on_closing():
        root.destroy()
        os._exit(0)  # 강제 종료

    root.protocol("WM_DELETE_WINDOW", on_closing)

# ====================================================================================================

# Python main loop & GUI
if __name__ == "__main__":

    # GUI 창 생성
    root = tk.Tk()
    root.title("SourceCode Analysis")
    root.geometry("300x400")

    # 창을 화면 중앙에 위치시키기
    window_width = 300
    window_height = 400
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int(screen_height / 2 - window_height / 2)
    position_right = int(screen_width / 2 - window_width / 2)
    root.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

    # 폴더 경로 변수 
    folder_path_var = tk.StringVar()
    output_folder_var = tk.StringVar()
    compile_commands_var = tk.StringVar()

    # 컴파일 커맨드 파일 경로 선택 및 경로 표시
    compile_commands_frame = tk.Frame(root)
    compile_commands_frame.pack(pady=5, padx=5, fill=tk.X)
    tk.Label(compile_commands_frame, text="compile_commands:").pack(side=tk.LEFT, padx=5, anchor='w')
    ttk.Button(compile_commands_frame, text="Browse", command=select_compile_commands, width=10).pack(side=tk.RIGHT, padx=5, anchor='e')
    ttk.Button(compile_commands_frame, text="Convert", command=convert_compile_commands, width=10).pack(side=tk.RIGHT, padx=5, anchor='e')
    ttk.Entry(root, textvariable=compile_commands_var).pack(pady=5, padx=5, fill=tk.X)

    # source code 폴더 선택 및 경로 표시
    input_frame = tk.Frame(root)
    input_frame.pack(pady=5, padx=5, fill=tk.X)
    tk.Label(input_frame, text="Source Code Folder:").pack(side=tk.LEFT, padx=5, anchor='w')
    ttk.Button(input_frame, text="Browse", command=select_input_folder, width=10).pack(side=tk.RIGHT, padx=5, anchor='e')
    ttk.Entry(root, textvariable=folder_path_var).pack(pady=5, padx=5, fill=tk.X)
    
    # output folder 선택 및 경로 표시
    output_frame = tk.Frame(root)
    output_frame.pack(pady=5, padx=5, fill=tk.X)
    tk.Label(output_frame, text="Output Folder:").pack(side=tk.LEFT, padx=5, anchor='w')
    ttk.Button(output_frame, text="Browse", command=select_output_folder, width=10).pack(side=tk.RIGHT, padx=5, anchor='e')
    ttk.Entry(root, textvariable=output_folder_var).pack(pady=5, padx=5, fill=tk.X)
    
    # 메인 함수 실행 버튼
    ttk.Button(root, text="Start", command=start_processing).pack(pady=5, padx=50, fill=tk.X)
    
    # 구분선
    separator = ttk.Separator(root, orient='horizontal')
    separator.pack(fill='x', pady=5, padx=10)
    separator.configure(style="TSeparator")

    # 진행 상황 표시
    progress_label = tk.Label(root, text="")
    progress_label.pack(pady=5)
    progress_bar = ttk.Progressbar(root, length=280, mode='determinate')
    progress_bar.pack(pady=5)
    file_label = tk.Label(root, text="")
    file_label.pack(pady=5)

    # output folder 열기 버튼
    open_button = ttk.Button(root, text="Open Output Folder", command=open_output_folder, state=tk.DISABLED)
    open_button.pack(pady=10)
    open_button.pack_forget()  # Hide the button initially

    API_button = ttk.Button(root, text="OpenAI API", command=api_call)
    API_button.pack(pady=10)

    root.mainloop()

    
