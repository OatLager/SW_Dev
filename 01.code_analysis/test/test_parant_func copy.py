import os
import tkinter as tk
from tkinter import filedialog
from collections import defaultdict
from clang.cindex import Index, CursorKind, Config
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import json
import shlex


# libclang 경로 설정 (시스템에 맞게 수정)
Config.set_library_file("C:/Program Files/LLVM/bin/libclang.dll")

DEFAULT_PROJECT_DIR = r'\\wsl.localhost\Ubuntu\home\jh\PX4-Autopilot\src\modules'
compile_commands_path = r"C:\Users\ypelec\Desktop\ing\방산100_FCS SW(박지훈)\작업 중\Code_v09\converted_compile_commands.json"

def select_directory_ui(title, initial_dir):
    root = tk.Tk()
    root.withdraw()
    return filedialog.askdirectory(title=title, initialdir=initial_dir)

def collect_cpp_files(folder):
    cpp_extensions = ['.c', '.cpp', '.cc', '.cxx', '.h', '.hpp', '.hh']
    return [os.path.join(root, file)
            for root, _, files in os.walk(folder)
            for file in files
            if os.path.splitext(file)[1] in cpp_extensions]

def get_clang_args(path_source_file, compile_commands_path=None, project_root=None):
    include_flags = []

    if compile_commands_path:
        print(f'컴파일 커맨드 경로 : {compile_commands_path}')
        try:
            with open(compile_commands_path, 'r') as f:
                compile_commands = json.load(f)
                for command in compile_commands:
                    try:
                        # 경로 정규화
                        command_path = os.path.normpath(command.get('file'))
                        target_path = os.path.normpath(path_source_file)

                        # 경로 비교 (같은 파일인지)
                        if os.path.samefile(command_path, target_path):
                            args = shlex.split(command['command'])
                            include_flags = [arg for arg in args if arg.startswith('-I') or arg.startswith('-isystem') or arg.startswith('-iquote')]
                            print(f"🔹 Found compile args for {os.path.basename(path_source_file)}")
                            return ["-x", "c++", "-std=c++17"] + include_flags
                    except FileNotFoundError:
                        # os.path.samefile이 파일이 없으면 에러를 낼 수 있어서 방어 코드
                        continue
        except Exception as e:
            print(f"⚠️ compile_commands.json 분석 실패: {e}")

    # fallback to manual args if compile_commands.json is missing
    if project_root:
        print(f'프로젝트 경로 : {project_root}')
        print("🔸 Falling back to default include paths")
        return [
            "-x", "c++", "-std=c++17",
            '-I' + project_root,
            '-I/usr/include',
            '-I/usr/include/c++/9',
            '-I/usr/include/x86_64-linux-gnu',
        ]
    else:
        return ["-x", "c++", "-std=c++17"]


# 함수 정의 및 호출 정보 수집
def find_functions_and_calls(file_path, compile_commands_path=None, project_root=None):
    index = Index.create()
    args = get_clang_args(file_path, compile_commands_path, project_root)

    try:
        tu = index.parse(file_path, args=args)
    except Exception as e:
        print(f"❌ {file_path} 파싱 실패: {e}")
        return [], []

    defined_funcs = []
    calls = []

    def get_context_name(node):
        context = []
        parent = node.semantic_parent
        while parent and parent.kind in [
            CursorKind.CLASS_DECL, CursorKind.STRUCT_DECL,
            CursorKind.UNION_DECL, CursorKind.NAMESPACE
        ]:
            if parent.spelling:
                context.insert(0, parent.spelling)
            parent = parent.semantic_parent
        context.append(node.spelling)
        return "::".join(context)

    def visitor(node, parent_func=None):
        if node.kind in [CursorKind.FUNCTION_DECL, CursorKind.CXX_METHOD, CursorKind.CONSTRUCTOR, CursorKind.DESTRUCTOR] and node.is_definition() and node.location.file and os.path.samefile(str(node.location.file), file_path):
            func_name = get_context_name(node)
            defined_funcs.append((func_name, file_path, node.location.line))  # 수정: line 정보 추가
            parent_func = func_name

        if node.kind == CursorKind.CALL_EXPR:
            called = node.displayname
            loc = node.location
            if loc.file:
                calls.append((called, loc.file.name, loc.line, parent_func))

        for c in node.get_children():
            visitor(c, parent_func)

    visitor(tu.cursor)
    return defined_funcs, calls

def analyze_files(files, compile_commands_path, project_root):
    defined_funcs = []
    calls = []
    for f in files:
        funcs, callsites = find_functions_and_calls(f, compile_commands_path, project_root)
        defined_funcs.extend(funcs)
        calls.extend(callsites)
    return defined_funcs, calls

def normalize_path(path):
    return os.path.normpath(path)

def get_relative_path(path, root):
    try:
        return os.path.relpath(path, root)
    except ValueError:
        return path

def sanitize_sheet_name(name):
    return name[:31]  # Excel 시트 이름은 최대 31자까지 허용

# 엑셀 저장 함수 수정
def save_to_excel(defined_funcs_map, matched_calls, output_path, base_filename, project_root):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for file_path, funcs in defined_funcs_map.items():
        sheet_name = sanitize_sheet_name(os.path.basename(file_path))  # 확장자 포함
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(['번호', '파일명', '함수 라인', '클래스::함수명', '호출자', '호출 위치', '호출자 경로'])

        for idx, (func_name, fpath, line_num) in enumerate(funcs, start=1):  # 수정: line_num 포함
            call_info = matched_calls.get(file_path, {}).get(func_name, [])
            if call_info:
                for (caller, caller_path, lines) in call_info:
                    loc_str = '\n'.join([f"Line {line}" for line in sorted(lines)])
                    rel_path = get_relative_path(caller_path, project_root)
                    sheet.append([
                        idx,
                        os.path.basename(file_path),
                        f"Line {line_num}",  # 수정: 함수 위치 = 라인
                        func_name,
                        caller or "",
                        loc_str,
                        rel_path or ""
                    ])
                    sheet.cell(row=sheet.max_row, column=6).alignment = Alignment(wrap_text=True)
            else:
                sheet.append([
                    idx,
                    os.path.basename(file_path),
                    f"Line {line_num}",
                    func_name,
                    "", "", ""
                ])

        for column in sheet.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0 for cell in column), default=10)
            col_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[col_letter].width = max_length + 5

    save_file = os.path.join(output_path, f"{base_filename}.xlsx")
    workbook.save(save_file)
    print(f"\n✅ 결과 저장 완료: {save_file}")

# 메인 함수에서도 func 리스트 포맷 변경 반영
if __name__ == "__main__":
    print("📁 전체 프로젝트 경로를 선택하세요.")
    project_root = select_directory_ui("전체 프로젝트 폴더 선택", DEFAULT_PROJECT_DIR)
    if not project_root:
        print("❌ 프로젝트 선택이 취소되었습니다.")
        exit()

    print("📁 분석할 대상 폴더를 선택하세요.")
    target_folder = select_directory_ui("분석할 대상 폴더 선택", project_root)
    if not target_folder:
        print("❌ 대상 폴더 선택이 취소되었습니다.")
        exit()

    print(f"📦 함수 정의 추출 중...")
    target_files = collect_cpp_files(target_folder)
    defined_funcs, _ = analyze_files(target_files, compile_commands_path, project_root)

    defined_funcs_map = defaultdict(list)
    for func, fpath, line in defined_funcs:  # 수정
        defined_funcs_map[normalize_path(fpath)].append((func, fpath, line))  # 수정

    print(f"✅ 정의된 함수 총 {len(defined_funcs)}개")

    print(f"\n🔍 전체 프로젝트에서 호출 위치 탐색 중...")
    project_files = collect_cpp_files(project_root)
    _, all_calls = analyze_files(project_files, compile_commands_path, project_root)
    print(f"📋 호출된 함수 총 {len(all_calls)}개:\n")

    matched_calls = defaultdict(lambda: defaultdict(list))

    for func_name, file_path, _ in defined_funcs:  # 수정
        base_func_name = func_name.split("::")[-1]
        for called_name, call_path, line, caller in all_calls:
            if called_name == base_func_name:
                norm_f = normalize_path(file_path)
                norm_cp = normalize_path(call_path)
                matched_calls[norm_f][func_name].append((caller, norm_cp, {line}))

    print("\n📥 저장할 폴더를 선택하세요.")
    save_dir = select_directory_ui("결과 저장 위치 선택", project_root)
    if save_dir:
        folder_name = os.path.basename(os.path.normpath(target_folder))
        save_to_excel(defined_funcs_map, matched_calls, save_dir, folder_name, project_root)
    else:
        print("❌ 저장 폴더 선택이 취소되었습니다.")